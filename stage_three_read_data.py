import openpyxl

def string_to_list_and_clean(string):  ## it reads the lists in the excel sheet as a string, this cleans it up as intended
    list1 = string.split()
    list2 = []
    for i in list1:
        temp = i
        temp = temp.translate({ord(c): None for c in """,'[]"""})
        list2.append(temp)
    return list2


def return_excel_data(cfd_output_path, project_name):

    workbook_path = f"{cfd_output_path}/{project_name} Variables.xlsx" # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 
    workbook = openpyxl.load_workbook(workbook_path) # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 

    worksheet = workbook.active

    Project_Name = worksheet["B1"].value
    Number_Of_Bedrooms = worksheet["B2"].value
    Number_Of_Lounges = worksheet["B3"].value
    Number_Of_Kitchens = worksheet["B4"].value
    Lounge_Fires = worksheet["B5"].value
    Lounge_Fires_Own_Door = string_to_list_and_clean(worksheet["B6"].value)
    Bedroom_Fires = worksheet["B7"].value
    Bedroom_Fires_Own_Door = string_to_list_and_clean(worksheet["B8"].value)
    Kitchen_Fires = worksheet["B9"].value
    Kitchen_Fires_Own_Door = string_to_list_and_clean(worksheet["B10"].value)
    TD_From_Bedrooms = string_to_list_and_clean(worksheet["B11"].value)
    TD_From_Kitchens = string_to_list_and_clean(worksheet["B12"].value)
    TD_From_Lounges = string_to_list_and_clean(worksheet["B13"].value)
    Suppression_Type = worksheet["B14"].value
    No_Scenarios = worksheet["B15"].value
    Scenario_Names = string_to_list_and_clean(worksheet["B16"].value)
    No_Openable_Doors = worksheet["B17"].value
    Default_Door = worksheet["B18"].value
    Scenario_Doors = worksheet["B19"].value
    Proposed_Detection = int(worksheet["B20"].value)
    CC_Detection = int(worksheet["B21"].value)
    Floor_To_Ceiling = worksheet["B22"].value
    
    return Number_Of_Bedrooms, Number_Of_Lounges, Number_Of_Kitchens, Kitchen_Fires, Suppression_Type, Lounge_Fires, Bedroom_Fires, No_Scenarios, Proposed_Detection, CC_Detection, Floor_To_Ceiling, Scenario_Names, Project_Name, TD_From_Bedrooms, TD_From_Lounges, TD_From_Kitchens


import pandas as pd
def return_scen_excel(path_to_results_file='CFD Test Output\Roneo Corner - Smallest Flat\Kitchen_Fire_1\Kitchen_Fire_1_Results.xlsx'):
    sheet_data = {}
    excel_data = {}
    workbook_path = f"{path_to_results_file}" # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 
    workbook = openpyxl.load_workbook(workbook_path, data_only=True) 
    for sheet in ['PD1', 'PD2', 'CC1', 'CC2', 'Event Trees']:
        sheet_data[sheet] = pd.read_excel(
                        path_to_results_file, 
                        index_col=0,
                        sheet_name=sheet,
                        )
    return sheet_data, workbook

# def return_scen_spreadsheet(path_to_scen='CFD Test Output\Roneo Corner - Smallest Flat\Kitchen_Fire_1'):

#     sheet_data = {}
#     for sheet in ['PD1', 'PD2', 'CC1', 'CC2', 'Event Trees']:
#         workbook_path = f"{path_to_scen}/{project_name} Variables.xlsx" # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 
#         workbook = openpyxl.load_workbook(workbook_path) # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 

#         worksheet = workbook.active