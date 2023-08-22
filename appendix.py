from constants import current_folder_path
from final_stage_functions import generate_data
from appendix_charts import prep_data, run_appendix_charts
import openpyxl
import json

# TODO: 100, 1000, 10000, 100000, 1mil *1k runs
# stage 3 but only the probability
def string_to_list_and_clean(string):  ## it reads the lists in the excel sheet as a string, this cleans it up as intended
    list1 = string.split()
    list2 = []
    for i in list1:
        temp = i
        temp = temp.translate({ord(c): None for c in """,'[]"""})
        list2.append(temp)
    return list2

from escape_probability import return_dataset_dict_from_curve
# TODO: take project name from path
def run_appendix(current_folder_path=current_folder_path, project_name='NHBC 8x10'):
    # pass
    Probability_Of_3m_Tenability_Limit = 0.3
    High_Visibility_Tenability_Limit = 3
    Low_Visibility_Tenability_Limit = 2

    Probability_Occupant_in_Kitchen = 0.03
    Probability_Occupant_in_Lounge = 0.46
    Probability_Occupant_in_Bedroom = 0.51  
    # Pre_dict = {14: 1, 15: 2, 16: 3, 17: 4, 18: 5, 19: 6, 20: 7, 21: 8, 22: 9, 23: 10, 24: 11, 25: 12, 26: 13, 27: 14, 28: 15, 29: 16, 30: 18, 31: 19, 32: 20, 33: 21, 34: 22, 35: 23, 36: 24, 37: 25, 38: 26, 39: 27, 40: 28, 41: 29, 42: 30, 43: 31, 44: 32, 45: 33, 46: 34, 47: 34, 48: 35, 49: 36, 50: 36, 51: 37, 52: 38, 53: 38, 54: 39, 55: 40, 56: 41, 57: 41, 58: 42, 59: 42, 60: 42, 61: 42, 62: 42, 63: 42, 64: 42, 65: 42, 66: 42, 67: 42, 68: 42, 69: 42, 70: 42, 71: 42, 72: 42, 73: 42, 74: 42, 75: 42, 76: 42, 77: 42, 78: 42, 79: 42, 80: 42, 81: 42, 82: 42, 83: 42, 84: 42, 85: 42, 86: 42, 87: 42, 88: 41, 89: 41, 90: 41, 91: 41, 92: 41, 93: 40, 94: 40, 95: 40, 96: 40, 97: 39, 98: 39, 99: 39, 100: 39, 101: 39, 102: 38, 103: 38, 104: 38, 105: 38, 106: 38, 107: 37, 108: 37, 109: 37, 110: 37, 111: 37, 112: 36, 113: 36, 114: 36, 115: 36, 116: 36, 117: 35, 118: 35, 119: 35, 120: 35, 121: 35, 122: 34, 123: 34, 124: 34, 125: 34, 126: 34, 127: 33, 128: 33, 129: 33, 130: 33, 131: 33, 132: 32, 133: 32, 134: 32, 135: 32, 136: 32, 137: 31, 138: 31, 139: 31, 140: 31, 141: 31, 142: 30, 143: 30, 144: 30, 145: 30, 146: 29, 147: 29, 148: 29, 149: 29, 150: 29, 151: 28, 152: 28, 153: 28, 154: 28, 155: 28, 156: 27, 157: 27, 158: 27, 159: 27, 160: 27, 161: 26, 162: 26, 163: 26, 164: 26, 165: 26, 166: 25, 167: 25, 168: 25, 169: 25, 170: 25, 171: 24, 172: 24, 173: 24, 174: 24, 175: 24, 176: 23, 177: 23, 178: 23, 179: 23, 180: 23, 181: 22, 182: 22, 183: 22, 184: 22, 185: 22, 186: 21, 187: 21, 188: 21, 189: 21, 190: 21, 191: 20, 192: 20, 193: 20, 194: 20, 195: 19, 196: 19, 197: 19, 198: 19, 199: 19, 200: 18, 201: 18, 202: 18, 203: 18, 204: 18, 205: 17, 206: 17, 207: 17, 208: 17, 209: 17, 210: 16, 211: 16, 212: 16, 213: 16, 214: 16, 215: 15, 216: 15, 217: 15, 218: 15, 219: 15, 220: 14, 221: 14, 222: 14, 223: 14, 224: 14, 225: 14, 226: 14, 227: 14, 228: 14, 229: 14, 230: 14, 231: 14, 232: 14, 233: 14, 234: 14, 235: 14, 236: 14, 237: 14, 238: 14, 239: 14, 240: 13, 241: 13, 242: 13, 243: 13, 244: 13, 245: 13, 246: 13, 247: 13, 248: 13, 249: 13, 250: 13, 251: 13, 252: 13, 253: 13, 254: 13, 255: 13, 256: 13, 257: 13, 258: 13, 259: 13, 260: 13, 261: 13, 262: 13, 263: 13, 264: 13, 265: 13, 266: 13, 267: 13, 268: 13, 269: 13, 270: 13, 271: 13, 272: 13, 273: 13, 274: 13, 275: 12, 276: 12, 277: 12, 278: 12, 279: 12, 280: 12, 281: 12, 282: 12, 283: 12, 284: 12, 285: 12, 286: 12, 287: 12, 288: 12, 289: 12, 290: 12, 291: 12, 292: 12, 293: 12, 294: 12, 295: 12, 296: 12, 297: 12, 298: 12, 299: 12, 300: 12, 301: 12, 302: 12, 303: 12, 304: 12, 305: 12, 306: 12, 307: 12, 308: 12, 309: 12, 310: 12, 311: 11, 312: 11, 313: 11, 314: 11, 315: 11, 316: 11, 317: 11, 318: 11, 319: 11, 320: 11, 321: 11, 322: 11, 323: 11, 324: 11, 325: 11, 326: 11, 327: 11, 328: 11, 329: 11, 330: 11, 331: 11, 332: 11, 333: 11, 334: 11, 335: 11, 336: 11, 337: 11, 338: 11, 339: 11, 340: 11, 341: 11, 342: 11, 343: 11, 344: 11, 345: 11, 346: 10, 347: 10, 348: 10, 349: 10, 350: 10, 351: 10, 352: 10, 353: 10, 354: 10, 355: 10, 356: 10, 357: 10, 358: 10, 359: 10, 360: 10, 361: 10, 362: 10, 363: 10, 364: 10, 365: 10, 366: 10, 367: 10, 368: 10, 369: 10, 370: 10, 371: 10, 372: 10, 373: 10, 374: 10, 375: 10, 376: 10, 377: 10, 378: 10, 379: 10, 380: 10, 381: 10, 382: 9, 383: 9, 384: 9, 385: 9, 386: 9, 387: 9, 388: 9, 389: 9, 390: 9, 391: 9, 392: 9, 393: 9, 394: 9, 395: 9, 396: 9, 397: 9, 398: 9, 399: 9, 400: 9, 401: 9, 402: 9, 403: 9, 404: 9, 405: 9, 406: 9, 407: 9, 408: 9, 409: 9, 410: 9, 411: 9, 412: 9, 413: 9, 414: 9, 415: 9, 416: 9, 417: 9, 418: 8, 419: 8, 420: 8, 421: 8, 422: 8, 423: 8, 424: 8, 425: 8, 426: 8, 427: 8, 428: 8, 429: 8, 430: 8, 431: 8, 432: 8, 433: 8, 434: 8, 435: 8, 436: 8, 437: 8, 438: 8, 439: 8, 440: 8, 441: 8, 442: 8, 443: 8, 444: 8, 445: 8, 446: 8, 447: 8, 448: 8, 449: 8, 450: 8, 451: 8, 452: 8, 453: 7, 454: 7, 455: 7, 456: 7, 457: 7, 458: 7, 459: 7, 460: 7, 461: 7, 462: 7, 463: 7, 464: 7, 465: 7, 466: 7, 467: 7, 468: 7, 469: 7, 470: 7, 471: 7, 472: 7, 473: 7, 474: 7, 475: 7, 476: 7, 477: 7, 478: 7, 479: 7, 480: 7, 481: 7, 482: 7, 483: 7, 484: 7, 485: 7, 486: 7, 487: 7, 488: 7, 489: 6, 490: 6, 491: 6, 492: 6, 493: 6, 494: 6, 495: 6, 496: 6, 497: 6, 498: 6, 499: 6, 500: 6, 501: 6, 502: 6, 503: 6, 504: 6, 505: 6, 506: 6, 507: 6, 508: 6, 509: 6, 510: 6, 511: 6, 512: 6, 513: 6, 514: 6, 515: 6, 516: 6, 517: 6, 518: 6, 519: 6, 520: 6, 521: 6, 522: 6, 523: 6, 524: 6, 525: 5, 526: 5, 527: 5, 528: 5, 529: 5, 530: 5, 531: 5, 532: 5, 533: 5, 534: 5, 535: 5, 536: 5, 537: 5, 538: 5, 539: 5, 540: 5, 541: 5, 542: 5, 543: 5, 544: 5, 545: 5, 546: 5, 547: 5, 548: 5, 549: 5, 550: 5, 551: 5, 552: 5, 553: 5, 554: 5, 555: 5, 556: 5, 557: 5, 558: 5, 559: 5, 560: 4, 561: 4, 562: 4, 563: 4, 564: 4, 565: 4, 566: 4, 567: 4, 568: 4, 569: 4, 570: 4, 571: 4, 572: 4, 573: 4, 574: 4, 575: 4, 576: 4, 577: 4, 578: 4, 579: 4, 580: 4, 581: 4, 582: 4, 583: 4, 584: 4, 585: 4, 586: 4, 587: 4, 588: 4, 589: 4, 590: 4, 591: 4, 592: 4, 593: 4, 594: 4, 595: 4, 596: 4, 597: 4, 598: 4, 599: 4, 600: 4, 601: 4, 602: 4, 603: 4, 604: 4, 605: 4, 606: 4, 607: 4, 608: 4, 609: 4, 610: 4, 611: 4, 612: 4, 613: 4, 614: 4, 615: 4, 616: 4, 617: 4, 618: 4, 619: 4, 620: 4, 621: 4, 622: 4, 623: 4, 624: 4, 625: 4, 626: 4, 627: 4, 628: 4, 629: 4, 630: 4, 631: 4, 632: 4, 633: 4, 634: 4, 635: 4, 636: 4, 637: 4, 638: 4, 639: 4, 640: 4, 641: 4, 642: 4, 643: 4, 644: 4, 645: 4, 646: 4, 647: 4, 648: 4, 649: 4, 650: 4, 651: 4, 652: 4, 653: 4, 654: 4, 655: 4, 656: 4, 657: 4, 658: 4, 659: 4, 660: 4, 661: 4, 662: 4, 663: 4, 664: 3, 665: 3, 666: 3, 667: 3, 668: 3, 669: 3, 670: 3, 671: 3, 672: 3, 673: 3, 674: 3, 675: 3, 676: 3, 677: 3, 678: 3, 679: 3, 680: 3, 681: 3, 682: 3, 683: 3, 684: 3, 685: 3, 686: 3, 687: 3, 688: 3, 689: 3, 690: 3, 691: 3, 692: 3, 693: 3, 694: 3, 695: 3, 696: 3, 697: 3, 698: 3, 699: 3, 700: 3, 701: 3, 702: 3, 703: 3, 704: 3, 705: 3, 706: 3, 707: 3, 708: 3, 709: 3, 710: 3, 711: 3, 712: 3, 713: 3, 714: 3, 715: 3, 716: 3, 717: 3, 718: 3, 719: 3, 720: 3, 721: 3, 722: 3, 723: 3, 724: 3, 725: 3, 726: 3, 727: 3, 728: 3, 729: 3, 730: 3, 731: 3, 732: 3, 733: 3, 734: 3, 735: 3, 736: 3, 737: 3, 738: 3, 739: 3, 740: 3, 741: 3, 742: 3, 743: 3, 744: 3, 745: 3, 746: 3, 747: 3, 748: 3, 749: 3, 750: 3, 751: 3, 752: 3, 753: 3, 754: 3, 755: 3, 756: 3, 757: 3, 758: 3, 759: 3, 760: 3, 761: 3, 762: 3, 763: 3, 764: 3, 765: 3, 766: 3, 767: 3, 768: 3, 769: 3, 770: 3, 771: 3, 772: 3, 773: 3, 774: 3, 775: 2, 776: 2, 777: 2, 778: 2, 779: 2, 780: 2, 781: 2, 782: 2, 783: 2, 784: 2, 785: 2, 786: 2, 787: 2, 788: 2, 789: 2, 790: 2, 791: 2, 792: 2, 793: 2, 794: 2, 795: 2, 796: 2, 797: 2, 798: 2, 799: 2, 800: 2, 801: 2, 802: 2, 803: 2, 804: 2, 805: 2, 806: 2, 807: 2, 808: 2, 809: 2, 810: 2, 811: 2, 812: 2, 813: 2, 814: 2, 815: 2, 816: 2, 817: 2, 818: 2, 819: 2, 820: 2, 821: 2, 822: 2, 823: 2, 824: 2, 825: 1, 826: 1, 827: 1, 828: 1, 829: 1, 830: 1, 831: 1, 832: 1, 833: 1, 834: 1, 835: 1, 836: 1, 837: 1, 838: 1, 839: 1, 840: 1, 841: 1, 842: 1, 843: 1, 844: 1, 845: 1, 846: 1, 847: 1, 848: 1, 849: 1, 850: 1, 851: 1, 852: 1, 853: 1, 854: 1, 855: 1, 856: 1, 857: 1, 858: 1, 859: 1, 860: 1, 861: 1, 862: 1, 863: 1, 864: 1, 865: 1, 866: 1, 867: 1, 868: 1, 869: 1, 870: 1, 871: 1, 872: 1, 873: 1}
    # ^ is for 10000 runs - need to devise a means of having this alter with the number of runs. 

    TC_From_Kitchens = ["1","2","3","4","5","6","7","8","9","10"]  ### a list of strings with thermocouple numbers
    TC_From_Lounges = ["1","2","3","4","5","6","7","8","9","10"]
    TC_From_Bedrooms = ["1","2","3","4","5","6","7","8","9"]

    Simulation_Time = 1200  ## need to get this from FDS model
    sprinkler_activation_time = 1000  ## programmed into FDS model - may need to be variable
    FED_Toxc_Tenability_Limit = 1.0  ### toxicity FED limit
    FED_Heat_Tenability_Limit_NS = 1.0  ### heat FED limit
    FED_RAD_tol = 1.33    #### radiation acceptable dose limit, as per 7974
    Temp_Tenability_Limit_S = 60  ## tenability limit for temperature should vapour be >10%
    Radiation_Tenability_Limit_S = 2.5 ### radiation tenability limit should vapour be >10%

    DensityCM = 1.14   # density of carbon monoxide kg/m3
    DensityCD = 1.87   # density of carbon dioxide kg/m3
    DCO = 30           # as per BS 7974:6
    VE = 25
    
    local_folder = current_folder_path
    current_folder_path = f'{current_folder_path}\Appendix Output'
    
    workbook_path = f"{current_folder_path}/{project_name}/{project_name} Variables.xlsx" # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 
    workbook = openpyxl.load_workbook(workbook_path) # create spreadsheet and worksheet. This will be on the drive so other parts of the program can read it. 

    worksheet = workbook.active

    Project_Name = worksheet["B1"].value
    Number_Of_Bedrooms = worksheet["B2"].value
    Number_Of_Lounges = worksheet["B3"].value
    Number_Of_Kitchens = worksheet["B4"].value
    Lounge_Fires = worksheet["B5"].value
    Lounge_Fires_Own_Door = string_to_list_and_clean(worksheet["B6"].value)
    Bedroom_Fires = worksheet["B7"].value
    Bedroom_Fires_Own_Door = string_to_list_and_clean(worksheet["B8"].value)
    Kitchen_Fires = worksheet["B9"].value
    Kitchen_Fires_Own_Door = string_to_list_and_clean(worksheet["B10"].value)
    TD_From_Bedrooms = string_to_list_and_clean(worksheet["B11"].value)
    TD_From_Kitchens = string_to_list_and_clean(worksheet["B12"].value)
    TD_From_Lounges = string_to_list_and_clean(worksheet["B13"].value)
    Suppression_Type = worksheet["B14"].value
    No_Scenarios = worksheet["B15"].value
    Scenario_Names = string_to_list_and_clean(worksheet["B16"].value)
    No_Openable_Doors = worksheet["B17"].value
    Default_Door = worksheet["B18"].value
    Scenario_Doors = worksheet["B19"].value
    Proposed_Detection = int(worksheet["B20"].value)
    CC_Detection = int(worksheet["B21"].value)
    Floor_To_Ceiling = worksheet["B22"].value


    detection_activation_value = 2


    def complete_runs(No_Runs):
        
        pre_move_dict = return_dataset_dict_from_curve(number_of_runs=No_Runs)
        Model_List = ["CC1", "CC2", "PD1", "PD2"]
        models_object = {}
        # n=0
        # while n < len(Scenario_Names): # for each scenario in the list of the scenario names
        results_dir = f"{current_folder_path}/{Project_Name}/{Scenario_Names[0]}"
        m=0
        #     file_name = f'{Scenario_Names[n]}_{num_runs}_Results.xlsx'

        #     try:    ##tries to gcreate spreadsheet
        #         dest_dir = f"{current_folder_path}/{Project_Name}/{Scenario_Names[n]}"  ### these lines copy the base model and paste it in the new model directory
        #         src_file = f"{local_folder}/template_results.xlsx"
        #         # shutil.copy(src_file, dest_dir)
        #         # os.rename(f"{current_folder_path}/{Project_Name}/{Scenario_Names[n]}/template_results.xlsx", f"{current_folder_path}/{Project_Name}/{Scenario_Names[n]}/{file_name}")
        #         # workbook = openpyxl.load_workbook(f"{current_folder_path}/{Project_Name}/{Scenario_Names[n]}/{file_name}") # creates spreadsheet for results
        #     except FileExistsError:
        #         print("file alaready exists - overwriting that file")

            # workbook = openpyxl.load_workbook(f"{current_folder_path}/{Project_Name}/{Scenario_Names[n]}/{file_name}") # creates spreadsheet for results
        while m < len(Model_List):  # for each model needed for each scenario - model list contains 4 models
            escape_fraction, trapped_fraction, harmed_fraction = generate_data(
                            f"{Model_List[m]}", 
                            f"{current_folder_path}/{Project_Name}/{Scenario_Names[0]}/{Model_List[m]}/{Model_List[m]}_devc.csv", 
                            f"{current_folder_path}/{Project_Name}/{Scenario_Names[0]}/{Model_List[m]}/{Model_List[m]}_hrr.csv", 
                            f"{Scenario_Names[0]}",
                            results_dir, 
                            Project_Name,
                            No_Runs,
                            FED_Toxc_Tenability_Limit,
                            DensityCM,
                            DCO,
                            FED_Heat_Tenability_Limit_NS,
                            Radiation_Tenability_Limit_S,
                            Temp_Tenability_Limit_S,
                            # workbook,
                            TC_From_Bedrooms,
                            pre_move_dict,
                            TD_From_Bedrooms,
                            Probability_Occupant_in_Bedroom,
                            Probability_Occupant_in_Lounge,
                            TD_From_Lounges,
                            TC_From_Lounges,
                            Probability_Occupant_in_Kitchen,
                            TD_From_Kitchens,
                            TC_From_Kitchens,
                            Probability_Of_3m_Tenability_Limit,
                            High_Visibility_Tenability_Limit,
                            Low_Visibility_Tenability_Limit,
                            detection_activation_value,
                            DensityCD,
                            FED_RAD_tol,
                            VE
                        )
                
            models_object[Model_List[m]] = {
                                            "escape_fraction": escape_fraction, 
                                            "trapped_fraction": trapped_fraction, 
                                            "harmed_fraction": harmed_fraction
                                            }
            m=m+1
            # TODO: add either further workbooks; or further sheets
        # return fractions from all 4 models - only PD and CC needed
        event_tree_obj = {}
        fire_remains = 0.23
        PD_no_sprinklers = fire_remains * 0.11
        event_tree_obj["PD"] = {
                                "trapped_fraction": PD_no_sprinklers*models_object["PD2"]["trapped_fraction"], 
                                "harmed_fraction": PD_no_sprinklers*models_object["PD2"]["harmed_fraction"]            
        }
        CC_door_open = fire_remains * 0.6
        CC_door_closed = fire_remains * 0.4
        event_tree_obj["CC"] = {
                                "trapped_fraction": CC_door_open*models_object["CC1"]["trapped_fraction"] + CC_door_closed*models_object["CC2"]["trapped_fraction"], 
                                "harmed_fraction": CC_door_open*models_object["CC1"]["harmed_fraction"] + CC_door_closed*models_object["CC2"]["harmed_fraction"]          
        }
        return event_tree_obj
        # do something num_runs; complete 1000 times
        # get average per num_runs

    
    run_list = {}
    stages = [100, 500, 1000, 5000, 10000, 50000, 100000]
    multiplier = 1000
    stages = [100, 500, 1000] # to be commented out
    multiplier = 100
    for runs_per_lap in stages: 
        run_list[runs_per_lap] = []
        for num_runs in [runs_per_lap for i in range(multiplier)]:
            models_object = complete_runs(num_runs)
            run_list[runs_per_lap].append(models_object)
            if len(run_list[runs_per_lap]) % 10 ==0:
                print(f'###{runs_per_lap}: {len(run_list[runs_per_lap])}###')
    # get average
        chart_data = prep_data(data=run_list)
        jsonString = json.dumps(chart_data)
        jsonFile = open(f"{project_name}_{runs_per_lap}x{multiplier}.json", "w")
        jsonFile.write(jsonString)
        jsonFile.close()
        pass
    # get the output
    # chart
    # TODO: save runs list to txt file
    # chart
    # TODO: run and save chart
    # run_appendix_charts(chart_data)
    pass
    
    # TODO: prep for modelling comp
    # TODO: should output file with PD, CC only

if __name__ == "__main__":
    # TODO: should queue both NHBC studies
    projects = ['NHBC 8x10', 'NHBC 12x16']
    for project in projects:
        run_appendix(project_name=project)
